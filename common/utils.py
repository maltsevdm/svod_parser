from tkinter import messagebox

from openpyxl.worksheet.worksheet import Worksheet

from common.consts import SVOD_ROW_COLUMNS, SVOD_ROW_HEADERS
from common.enums import SvodHeaders
from common.exceptions import ColumnNotFound


def get_svod_columns(ws: Worksheet) -> dict[str, int]:
    headers = []
    for col in range(1, ws.max_column + 1):
        value = str(ws.cell(SVOD_ROW_HEADERS, col).value).strip()

        headers.append(value)

    columns = {}
    for header in SvodHeaders:
        h_value = header.value
        if header.name.startswith("ЕД_ИЗМ"):
            h_value = h_value[:-1]

        if h_value not in headers:
            raise ColumnNotFound(f"Не удалось найти колонку {header}")

        h_index = headers.index(h_value)
        columns[header] = h_index + 1
        if header.name.startswith("ЕД_ИЗМ"):
            headers[h_index] = "Ед. изм. temp."

    return columns


def find_flat_column_index(ws: Worksheet, flat_column: int) -> int | None:
    for col in range(1, ws.max_column + 1):
        if ws.cell(SVOD_ROW_COLUMNS, col).value == flat_column:
            return col
    messagebox.showerror(
        "Ошибка", f"В строке {SVOD_ROW_COLUMNS} не найдена колонка {flat_column}"
    )
