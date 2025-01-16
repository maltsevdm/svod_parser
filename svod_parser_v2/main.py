import logging
import os
import pathlib
from tkinter import filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.cell import get_column_letter
from openpyxl_image_loader import SheetImageLoader

from common.consts import (
    flat_column,
    SVOD_ROW_FLATS,
    SVOD_SHEET,
    svod_static_rows,
)
from common.enums import SvodHeaders
from common.exceptions import ColumnNotFound
from common.styles import GREEN_COLOR, thin_border
from common.utils import find_flat_column_index, get_svod_columns

from .consts import (
    IMAGE_MAX_HEIGHT,
    IMAGE_MAX_WIDTH,
    ROW_MAIN_HEIGHT,
    ROW_MINOR_HEIGHT,
    TEMPLATE_FILENAME,
    columns_relation,
)
from .enums import SpecColumn
from .utils import transform_formula

current_dir = os.path.dirname(os.path.abspath(__file__))
template_path = os.path.join(current_dir, TEMPLATE_FILENAME)


def start_process():
    file_path = pathlib.Path(
        filedialog.askopenfilename(
            title="Выберите файл Свод", filetypes=[("Excel files", "*.xlsx")]
        )
    )
    try:
        if main(flat_column.get(), file_path):
            messagebox.showinfo("Информация", "Готово")
    except Exception as ex:
        logging.exception(ex)
        messagebox.showerror("Ошибка", "Произошла непредвиденная ошибка")


def main(flat_column: int, file_path: pathlib.Path) -> bool:
    wb_source_with_formulas = load_workbook(file_path)
    try:
        ws_svod = wb_source_with_formulas[SVOD_SHEET]
    except KeyError:
        messagebox.showerror(
            "Ошибка",
            f'В загруженном файле нет листа "{SVOD_SHEET}", я умею работать только с таким листом',
        )
        return

    wb_src = load_workbook(file_path, data_only=True)
    ws_svod_do = wb_src[SVOD_SHEET]

    wb_new = load_workbook(template_path)
    ws_spec = wb_new.active

    sc = get_svod_columns(ws_svod)

    flat_column_i = find_flat_column_index(ws_svod, flat_column)
    if not flat_column_i:
        return False

    flat = ws_svod.cell(SVOD_ROW_FLATS, flat_column_i).value

    ws_spec.title = str(flat)

    row_spec = 8

    i = 1

    try:
        image_loader = SheetImageLoader(ws_svod)
    except IndexError as ex:
        if "string index out of range" in str(ex):
            messagebox.showerror(
                "Ошибка",
                "В своде скорее всего есть картинки в столбцах кроме Внешний вид. "
                "Удалите их и попробуйте снова.",
            )
            return
        raise ex

    # Переносим данные
    for row in range(9, ws_svod.max_row + 1):
        if str(ws_svod.cell(row, 3).value).lower() in svod_static_rows:
            ws_spec.cell(
                row_spec,
                SpecColumn.НОМЕР,
                ws_svod.cell(row, 3).value.upper(),
            )
            row_spec += 1
            continue

        if not ws_svod.cell(row, flat_column_i).value:
            continue
        
        cell_address = f"{get_column_letter(sc[SvodHeaders.ВНЕШНИЙ_ВИД])}{row}"

        if image_loader.image_in(cell_address):
            im = Image(image_loader.get(cell_address))

            if im.width < IMAGE_MAX_WIDTH:
                im.height = IMAGE_MAX_WIDTH * im.height / im.width
                im.width = IMAGE_MAX_WIDTH

            if im.height > IMAGE_MAX_HEIGHT:
                im.width = IMAGE_MAX_HEIGHT * im.width / im.height
                im.height = IMAGE_MAX_HEIGHT

            if im.width > IMAGE_MAX_WIDTH:
                im.height = IMAGE_MAX_WIDTH * im.height / im.width
                im.width = IMAGE_MAX_WIDTH

            cell_address = f"{get_column_letter(SpecColumn.ВНЕШНИЙ_ВИД)}{row_spec}"
            ws_spec.add_image(im, cell_address)

        ws_spec.cell(row_spec, SpecColumn.НОМЕР, i)

        for col_from_h, col_to in columns_relation.items():
            if col_from_h == SvodHeaders.СТРОИТЕЛЬНЫЙ_ЗАПАС_ПРОЦ:
                continue

            if col_from_h in [
                SvodHeaders.ЦЕНА_ОТ_ЗАКУПОК_ОПТ,
                SvodHeaders.ЦЕНА_ОТ_ЗАКУПОК_РОЗН,
            ]:
                ws_sv = ws_svod_do
            else:
                ws_sv = ws_svod

            col_from = sc[col_from_h]
            value = ws_sv.cell(row, col_from).value
            if str(value).startswith("="):
                try:
                    new_formula = transform_formula(value, sc).format(row=row_spec)
                except ColumnNotFound:
                    address = f"{get_column_letter(col_from)}{row}"
                    messagebox.showerror(
                        "Ошибка",
                        f"В своде в ячейке {address} в формуле используется колонка, которой нет в итоговом файле",
                    )
                    return

                ws_spec.cell(row_spec, col_to, new_formula)
            else:
                ws_spec.cell(row_spec, col_to, ws_sv.cell(row, col_from).value)

        row_spec += 1
        i += 1

    # Делаем оформление
    for row in range(8, ws_spec.max_row + 1):
        if not ws_spec.cell(row, 1).value:
            break

        if str(ws_spec.cell(row, SpecColumn.НОМЕР).value).lower() in svod_static_rows:
            ws_spec.row_dimensions[row].height = ROW_MINOR_HEIGHT
            ws_spec.merge_cells(
                start_row=row,
                start_column=SpecColumn.НОМЕР,
                end_row=row,
                end_column=SpecColumn.ПОЛНАЯ_КОМПЛЕКТАЦИЯ + 1,
            )
            cell = ws_spec.cell(row, SpecColumn.НОМЕР)
            cell.fill = PatternFill("solid", GREEN_COLOR)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            ws_spec.border = thin_border
            continue

        ws_spec.row_dimensions[row].height = ROW_MAIN_HEIGHT

        for col in range(1, SpecColumn.ПОЛНАЯ_КОМПЛЕКТАЦИЯ + 2):
            cell = ws_spec.cell(row, col)

            cell.border = thin_border

            if col in [
                SpecColumn.ОБЪЕМЫ_С_РП,
                SpecColumn.ОБЪЕМ_ДЛЯ_ЗАКУПА,
                SpecColumn.ИТОГО_ОБЪЕМ_С_ЗАПАСОМ_ДЛЯ_ЗАКУПА,
            ]:
                ws_spec.cell(row, col).number_format = "0.0"

            if col in [
                SpecColumn.ОЖИДАЕМАЯ_СКИДКА_К_ЦЕНЕ,
                SpecColumn.СТРОИТЕЛЬНЫЙ_ЗАПАС_ПРОЦ,
            ]:
                ws_spec.cell(row, col).number_format = "0%"

            if col in [
                SpecColumn.ЦЕНА_ОТ_ЗАКУПОК_ОПТ,
                SpecColumn.ЦЕНА_ОТ_ЗАКУПОК_РОЗН,
                SpecColumn.ИТОГО_ЦЕНА,
                SpecColumn.ИТОГО_СТОИМОСТЬ_ЧИСТОВОЙ_ОТДЕЛКИ,
                SpecColumn.ЧИСТОВАЯ_И_КОРПУСНАЯ_МЕБЕЛЬ,
                SpecColumn.ПОЛНАЯ_КОМПЛЕКТАЦИЯ,
            ]:
                ws_spec.cell(row, col).number_format = "0"

            if col - 1 in [
                SpecColumn.ПОСТАВЩИК,
                SpecColumn.ИТОГО_ЦЕНА,
                SpecColumn.ОБЪЕМЫ_С_РП,
                SpecColumn.ИТОГО_ОБЪЕМ_С_ЗАПАСОМ_ДЛЯ_ЗАКУПА,
                SpecColumn.ПОЛНАЯ_КОМПЛЕКТАЦИЯ,
            ]:
                cell.fill = PatternFill("solid", GREEN_COLOR)

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    try:
        wb_new.save(file_path.with_name("Excel комплектация квартира.xlsx"))
    except PermissionError:
        messagebox.showerror(
            "Ошибка",
            "Закройте файл Excel комплектация квартира.xlsx и попробуйте снова",
        )
        return False

    return True
