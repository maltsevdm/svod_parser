import logging
import os
import pathlib
from tkinter import filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl_image_loader import SheetImageLoader

from common.enums import SvodHeaders
from common.styles import GREEN_COLOR, thin_border
from common.utils import get_svod_columns

from .consts import (
    COL_L_FORMULA,
    IMAGE_MAX_HEIGHT,
    IMAGE_MAX_WIDTH,
    ROW_COLUMNS,
    ROW_FLATS,
    ROW_MAIN_HEIGHT,
    ROW_MINOR_HEIGHT,
    ROW_NAMES,
    SVOD_SHEET,
    TEMPLATE_FILENAME,
    columns_relation,
    flat_column,
    static_rows,
)
from .enums import SpecColumn
from .exceptions import ColumnNotFound
from .utils import transform_formula

current_dir = os.path.dirname(os.path.abspath(__file__))
template_path = os.path.join(current_dir, TEMPLATE_FILENAME)


def start_process():
    file_path = pathlib.Path(
        filedialog.askopenfilename(
            title="Выберите файл Свод", filetypes=[("Excel files", "*.xlsx")]
        )
    )
    try:
        if main(flat_column.get(), file_path):
            messagebox.showinfo("Информация", "Готово")
    except Exception as ex:
        logging.exception(ex)
        messagebox.showerror("Ошибка", "Произошла непредвиденная ошибка")


def find_flat_column_index(ws: Worksheet, flat_column: int) -> int | None:
    for col in range(1, ws.max_column + 1):
        if ws.cell(ROW_COLUMNS, col).value == flat_column:
            return col
    messagebox.showerror(
        "Ошибка", f"В строке {ROW_COLUMNS} не найдена колонка {flat_column}"
    )


def main(flat_column: int, file_path: pathlib.Path) -> bool:
    wb_source_with_formulas = load_workbook(file_path)
    try:
        ws_svod_with_formulas = wb_source_with_formulas[SVOD_SHEET]
    except KeyError:
        messagebox.showerror(
            "Ошибка",
            f'В загруженном файле нет листа "{SVOD_SHEET}", я умею работать только с таким листом',
        )
        return

    wb_source = load_workbook(file_path, data_only=True)
    ws_svod = wb_source[SVOD_SHEET]

    wb_new = load_workbook(template_path)
    ws_spec = wb_new.active

    sc = get_svod_columns(ws_svod)

    flat_column_i = find_flat_column_index(ws_svod, flat_column)
    if not flat_column_i:
        return False

    flat = ws_svod.cell(ROW_FLATS, flat_column_i).value

    ws_spec.title = str(flat)

    row_spec = 3
    # Квартира
    ws_spec.cell(1, SpecColumn.КОЛВО_ДЛЯ_ЗАКУПА_С_РП, flat)
    # Имя
    ws_spec.cell(
        1,
        SpecColumn.КОЛВО_ДЛЯ_ЗАКУПА_С_РП + 1,
        ws_svod.cell(ROW_NAMES, flat_column_i).value,
    )

    i = 1

    
    try:
        image_loader = SheetImageLoader(ws_svod)
    except IndexError as ex:
        if "string index out of range" in str(ex):
            messagebox.showerror(
                "Ошибка",
                "В своде скорее всего есть картинки в столбцах кроме Внешний вид. "
                "Удалите их и попробуйте снова.",
            )
            return
        raise ex

    # Переносим данные
    for row in range(ROW_FLATS + 1, ws_svod.max_row + 1):
        col_static = sc[SvodHeaders.НАИМЕНОВАНИЕ_ПО_ПРОЕКТУ]
        if str(ws_svod.cell(row, col_static).value).lower() in static_rows:
            ws_spec.cell(
                row_spec,
                1,
                ws_svod.cell(row, col_static).value.upper(),
            )
            row_spec += 1
            continue

        if not ws_svod.cell(row, flat_column_i).value:
            continue

        cell_address = f"{get_column_letter(sc[SvodHeaders.ВНЕШНИЙ_ВИД])}{row}"

        if image_loader.image_in(cell_address):
            im = Image(image_loader.get(cell_address))

            if im.width < IMAGE_MAX_WIDTH:
                im.height = IMAGE_MAX_WIDTH * im.height / im.width
                im.width = IMAGE_MAX_WIDTH

            if im.height > IMAGE_MAX_HEIGHT:
                im.width = IMAGE_MAX_HEIGHT * im.width / im.height
                im.height = IMAGE_MAX_HEIGHT

            if im.width > IMAGE_MAX_WIDTH:
                im.height = IMAGE_MAX_WIDTH * im.height / im.width
                im.width = IMAGE_MAX_WIDTH

            cell_address = f"{get_column_letter(SpecColumn.ВНЕШНИЙ_ВИД)}{row_spec}"
            ws_spec.add_image(im, cell_address)

        ws_spec.cell(row_spec, SpecColumn.НОМЕР, i)

        for col_from, col_to in columns_relation.items():
            ws_spec.cell(
                row_spec,
                col_to,
                ws_svod.cell(row, sc[col_from]).value,
            )

        for col_with_formula in [
            SvodHeaders.КОЛВО_ДЛЯ_ЗАКУПА_С_УЧЕТОМ_ЗАПАСА,
            SvodHeaders.КОЛВО_ДЛЯ_ЗАКУПА_С_РП,
        ]:
            value = ws_svod_with_formulas.cell(row, sc[col_with_formula]).value
            if str(value).startswith("="):
                try:
                    new_formula = transform_formula(value, sc).format(row=row_spec)
                except ColumnNotFound:
                    address = f"{get_column_letter(sc[col_with_formula])}{row}"
                    messagebox.showerror(
                        "Ошибка",
                        f"В своде в ячейке {address} в формуле используется колонка, которой нет в итоговом файле",
                    )
                    return

                ws_spec.cell(
                    row_spec,
                    columns_relation[col_with_formula],
                    new_formula,
                )

        ws_spec.cell(
            row_spec,
            SpecColumn.КОЛВО_ОБЩЕЕ_ДЛЯ_ЗАКУПА,
            COL_L_FORMULA.format(row=row_spec),
        )

        if ws_svod.cell(row, flat_column_i).fill != PatternFill(None):
            ws_spec.cell(
                row_spec,
                SpecColumn.МАТЕРИАЛЫ_С_РП,
                ws_svod.cell(row, flat_column_i).value,
            )

        row_spec += 1
        i += 1

    # Делаем оформление
    for row in range(3, ws_spec.max_row + 1):
        if not ws_spec.cell(row, 1).value:
            break

        if str(ws_spec.cell(row, 1).value).lower() in static_rows:
            ws_spec.row_dimensions[row].height = ROW_MINOR_HEIGHT
            ws_spec.merge_cells(
                start_row=row,
                start_column=1,
                end_row=row,
                end_column=21,
            )
            cell = ws_spec.cell(row, 1)
            cell.fill = PatternFill("solid", GREEN_COLOR)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            ws_spec.border = thin_border
            continue

        if row >= 3:
            ws_spec.row_dimensions[row].height = ROW_MAIN_HEIGHT
            ws_spec.cell(row, SpecColumn.СТРОИТЕЛЬНЫЙ_ЗАПАС).number_format = "0%"

        for col in range(1, SpecColumn.МАТЕРИАЛЫ_ДЛЯ_ЗАКУПА_С_РП + 2):
            cell = ws_spec.cell(row, col)

            # Делаем зеленые разделительные границы
            if row >= 2 and col - 1 in [
                SpecColumn.НАИМЕНОВАНИЕ_ПО_РП,
                SpecColumn.ЦЕНА_С_НДС,
                SpecColumn.ПОСТАВЩИК,
                SpecColumn.ЕД_ИЗМ_1,
                SpecColumn.СТРОИТЕЛЬНЫЙ_ЗАПАС,
                SpecColumn.МАТЕРИАЛЫ_С_РП,
                SpecColumn.МАТЕРИАЛЫ_ДЛЯ_ЗАКУПА_С_РП,
            ]:
                cell.fill = PatternFill("solid", GREEN_COLOR)

            if row >= 3:
                cell.border = thin_border

            if col in [
                SpecColumn.КОЛВО_ОБЩЕЕ_ДЛЯ_ЗАКУПА,
                SpecColumn.КОЛВО_ДЛЯ_ЗАКУПА_С_РП,
                SpecColumn.МАТЕРИАЛЫ_С_РП,
                SpecColumn.МАТЕРИАЛЫ_ДЛЯ_ЗАКУПА_С_РП,
            ]:
                ws_spec.cell(row, col).number_format = "0.0"

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    try:
        wb_new.save(file_path.with_name(f"{flat}.xlsx"))
    except PermissionError:
        messagebox.showerror("Ошибка", f"Закройте файл {flat}.xlsx и попробуйте снова")
        return False

    return True
